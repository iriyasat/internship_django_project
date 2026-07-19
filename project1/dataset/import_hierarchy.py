import os
import sys
import django
import openpyxl

# Set up Django environment
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'project1.settings')
django.setup()

from car_sales.models import EmployeeHierarchy

def import_hierarchy():
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'employee_hierarchy_sheet.xlsx')
    print(f"Opening Excel file: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if 'employee_hierarchy' not in wb.sheetnames:
        print("Error: Sheet 'employee_hierarchy' not found!")
        return
        
    ws = wb['employee_hierarchy']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]
    
    # Filter for valid rows (employee_id and role_id must be integers)
    valid_rows = []
    for row in data_rows:
        if isinstance(row[0], int) and isinstance(row[1], int):
            valid_rows.append(row)
            
    print(f"Found {len(valid_rows)} valid hierarchy records.")
    
    # Clear existing hierarchy records
    print("Clearing existing employee hierarchies...")
    EmployeeHierarchy.objects.all().delete()
    
    # Prepare bulk create
    hierarchies = []
    for r in valid_rows:
        hierarchies.append(EmployeeHierarchy(
            employee_id=r[0],
            role_id=r[1],
            supervisor_id=r[2],
            supervisor_role_id=r[3],
            supervisor2_id=r[4],
            supervisor2_role_id=r[5],
            supervisor3_id=r[6],
            supervisor3_role_id=r[7],
            supervisor4_id=r[8],
            supervisor4_role_id=r[9],
            supervisor5_id=r[10],
            supervisor5_role_id=r[11]
        ))
        
    print("Saving to database in batches...")
    EmployeeHierarchy.objects.bulk_create(hierarchies, batch_size=500)
    print("Import completed successfully!")

if __name__ == '__main__':
    import_hierarchy()
