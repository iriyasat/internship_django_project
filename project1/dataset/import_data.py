import os
import time
import django
import openpyxl

# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'project1.settings')
django.setup()

from django.db import connection

excel_path = os.path.join('dataset', 'car_sales_dataset_v2_untouched.xlsx')

SHEETS_CONFIG = [
    {
        'name': 'country',
        'table': 'country',
        'fields': 'country_id, country_name, created_at, updated_at',
        'row_placeholders': '(%s, %s, NOW(6), NOW(6))',
        'cols_count': 2
    },
    {
        'name': 'city',
        'table': 'city',
        'fields': 'city_id, city_name, country_id, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 3
    },
    {
        'name': 'store',
        'table': 'store',
        'fields': 'store_id, store_name, store_code, city_id, country_id, address, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 6
    },
    {
        'name': 'employee_role',
        'table': 'employee_role',
        'fields': 'role_id, role_name, created_at, updated_at',
        'row_placeholders': '(%s, %s, NOW(6), NOW(6))',
        'cols_count': 2
    },
    {
        'name': 'employee_status',
        'table': 'employee_status',
        'fields': 'status_id, status, created_at, updated_at',
        'row_placeholders': '(%s, %s, NOW(6), NOW(6))',
        'cols_count': 2
    },
    {
        'name': 'employee',
        'table': 'employee',
        'fields': 'employee_id, first_name, last_name, date_of_joining, employee_addr, employee_role, status, store_id, city_id, country_id, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 10
    },
    {
        'name': 'industry_info',
        'table': 'industry_info',
        'fields': 'make_id, make_name, created_at, updated_at',
        'row_placeholders': '(%s, %s, NOW(6), NOW(6))',
        'cols_count': 2
    },
    {
        'name': 'vehicle_info',
        'table': 'vehicle_info',
        'fields': 'id, vehicle_model, make_id, mmr, trim, body, transmission, vin, state, `condition`, odometer, color, interior, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 13
    },
    {
        'name': 'customer_info',
        'table': 'customer_info',
        'fields': 'customer_id, firstname, lastname, customer_status, customer_address, city_id, country_id, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 7
    },
    {
        'name': 'selling_info',
        'table': 'selling_info',
        'fields': 'sell_id, customer_id, vehicle_id, employee_id, store_id, selling_price, selling_date, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 7
    },
    {
        'name': 'employee_budget',
        'table': 'employee_budget',
        'fields': 'employee_id, budget_year, budget_month, store_id, budget_qty, budget_amount, created_at, updated_at',
        'row_placeholders': '(%s, %s, %s, %s, %s, %s, NOW(6), NOW(6))',
        'cols_count': 6
    }
]

def main():
    start_total_time = time.time()
    print("Starting optimized database clear and Excel data import...")
    
    with connection.cursor() as cursor:
        print("Disabling foreign key checks and truncating tables...")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
        for config in reversed(SHEETS_CONFIG):
            table = config['table']
            print(f"Truncating table: {table}")
            cursor.execute(f"TRUNCATE TABLE {table};")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
        print("Tables cleared successfully.")

    print(f"Opening Excel file: {excel_path} (this might take a few seconds)...")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    
    # 5,000 - 10,000 is an optimal batch size for MySQL single queries
    BATCH_SIZE = 5000
    
    with connection.cursor() as cursor:
        print("Disabling foreign key checks for import...")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
        
        for config in SHEETS_CONFIG:
            sheet_name = config['name']
            table = config['table']
            fields = config['fields']
            row_placeholder = config['row_placeholders']
            cols_count = config['cols_count']
            
            print(f"\nProcessing sheet '{sheet_name}' -> table '{table}'")
            if sheet_name not in wb.sheetnames:
                print(f"Warning: Sheet {sheet_name} not found in workbook. Skipping.")
                continue
                
            sheet = wb[sheet_name]
            rows_gen = sheet.iter_rows(values_only=True)
            
            # Skip header row
            try:
                next(rows_gen)
            except StopIteration:
                print(f"Sheet {sheet_name} is empty.")
                continue
            
            batch = []
            count = 0
            start_sheet_time = time.time()
            
            def flush_batch(b):
                if not b:
                    return
                # Build a single multi-row insert query
                placeholders = ", ".join([row_placeholder] * len(b))
                sql = f"INSERT INTO {table} ({fields}) VALUES {placeholders}"
                flat_args = [val for row in b for val in row]
                cursor.execute(sql, flat_args)
            
            for row in rows_gen:
                if not row or row[0] is None:
                    # Skip empty rows or rows with missing ID
                    continue
                
                # Grab only the expected number of columns
                row_data = row[:cols_count]
                batch.append(row_data)
                count += 1
                
                if len(batch) >= BATCH_SIZE:
                    flush_batch(batch)
                    batch = []
                    print(f"  Inserted {count} rows...")
            
            # Insert remaining rows
            if batch:
                flush_batch(batch)
            
            elapsed = time.time() - start_sheet_time
            print(f"Completed sheet '{sheet_name}': {count} rows inserted in {elapsed:.2f} seconds.")
            
        print("\nRe-enabling foreign key checks...")
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
        
    total_elapsed = time.time() - start_total_time
    print(f"\nAll operations completed successfully in {total_elapsed:.2f} seconds.")

if __name__ == '__main__':
    main()
